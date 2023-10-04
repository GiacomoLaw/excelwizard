import win32com.client
import os
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# TODO: paste numbers as numbers - currently the numbers are pasted as text, which causes issues with formula on other sheets
# TODO: dynamic sheet insertion depending on attachment file name (sales vs quotes)

# outlook object creation
outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

# search through inbox
folder = outlook.GetDefaultFolder(6)

# get latest items
emails = folder.Items
emails.Sort("[ReceivedTime]", True)

for email in emails:
    # define email subject to search for
    if email.Subject == "CSV file from Report Wizard":
        subject = email.Subject

        # attachment check
        if email.Attachments.Count > 0:
            # save to temp file
            attachment = email.Attachments.Item(1)
            attachment_filename = os.path.join(os.getcwd(), attachment.FileName)
            attachment.SaveAsFile(attachment_filename)
            print(f"Saved attachment: {attachment_filename}")

            # add to dataframe with windows encoding
            df = pd.read_csv(attachment_filename, encoding='cp1252')

            # Convert dates and change date formats
            df['Date Entered'] = pd.to_datetime(df['Date Entered'], format='%d/%m/%y').dt.strftime('%d/%m/%Y')
            df['Date Promised'] = pd.to_datetime(df['Date Promised'], format='%d/%m/%y').dt.strftime('%d/%m/%Y')
            df['Date Despatched'] = pd.to_datetime(df['Date Despatched'], format='%d/%m/%y').dt.strftime('%d/%m/%Y')

            # Sort criteria
            df.sort_values(by=['Order No', 'Order Line', 'Date Entered'], inplace=True)

            # Format the 'Net Value' and 'Customer Group' columns as numbers
            df['Net Value'] = df['Net Value'].str.replace(',', '', regex=True).astype(float)
            df['Customer Group'] = pd.to_numeric(df['Customer Group'], errors='coerce')  # Handle non-numeric values

            df.to_clipboard(index=False, header=True, sep='\t', decimal=',')

            print(f'Sorted and copied.')

            excel_file_path = 'C:\\Users\\marketing\\Documents\\2023 SALES FOLLOW UP.xlsx'
            wb = openpyxl.load_workbook(excel_file_path)
            print('File opened.')

            sheet_name = 'Sales Order Data'
            ws = wb[sheet_name]
            print('Sheet selected.')

            # clear existing data in the sheet, apart from columns T, U, and V
            for row_index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=19), 2):
                for cell_index, cell in enumerate(row, 1):
                    if cell.column_letter not in ['T', 'U', 'V']:
                        cell.value = None
            print('Data cleared, inserting.')

            # insert sorted data
            for row_index, row_data in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for column_index, value in enumerate(row_data, 1):
                    ws.cell(row=row_index, column=column_index, value=value)

            wb.save(excel_file_path)
            wb.close()

            os.remove(attachment_filename)
            print('Inserted and cleaned. Complete. B)')
        else:
            print("No attachments found in the email.")

        break

# release outlook
del outlook
