import win32com.client
import os
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# TODO: paste numbers as numbers - currently the numbers are pasted as text, which causes issues with formula on other sheets
# TODO: prevent first row (headers) being copied on spreadsheet for cleaner insertion
# TODO: dynamic sheet insertion depending on attachment file name (sales vs quotes)

# outlook object creation
outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

# search through inbox
folder = outlook.GetDefaultFolder(6)

# get latest items
emails = folder.Items
emails.Sort("[ReceivedTime]", True)

for email in emails:
    # define email subject to search for
    if email.Subject == "CSV file from Report Wizard":
        subject = email.Subject
        print(f"Subject: {subject}")

        # attachment check
        if email.Attachments.Count > 0:
            # save to temp file
            attachment = email.Attachments.Item(1)
            attachment_filename = os.path.join(os.getcwd(), attachment.FileName)
            attachment.SaveAsFile(attachment_filename)
            print(f"Saved attachment: {attachment_filename}")

            # add to dataframe with windows encoding
            df = pd.read_csv(attachment_filename, encoding='cp1252')

            # sort criteria
            df.sort_values(by=['Order No', 'Order Line', 'Date Entered'], inplace=True)

            # copy dataframe to clipboard
            df.to_clipboard(index=False, header=True, sep='\t')

            print(f'Successfully sorted and copied to the clipboard with columns.')

            # open existing excel file
            excel_file_path = 'C:\\Users\\marketing\\Documents\\2023 SALES FOLLOW UP.xlsx'
            wb = openpyxl.load_workbook(excel_file_path)
            
            # specify sheet
            sheet_name = 'Sales Order Data'
            ws = wb[sheet_name]

            # clear existing data apart from first row and later columns
            for row_index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=19), 2):
                for cell_index, cell in enumerate(row, 1):
                    if cell.column_letter not in ['T', 'U', 'V']:
                        cell.value = None

            # insert sorted data
            for row_index, row_data in enumerate(dataframe_to_rows(df, index=False, header=True), 2):
                for column_index, value in enumerate(row_data, 1):
                    ws.cell(row=row_index, column=column_index, value=value)

            wb.save(excel_file_path)
            wb.close()

            os.remove(attachment_filename)
        else:
            print("No attachments found in the email.")

        break

# release outlook
del outlook
